#!/usr/bin/env python3

from typing import List
from tqdm import tqdm
from selenium import webdriver
from webdriver_manager import manager

from webdriver_manager.firefox import GeckoDriverManager
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.microsoft import IEDriverManager

from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from openpyxl import Workbook,load_workbook
from datetime import datetime
import argparse
import os
import sys
import time
import re
import pathlib
import csv
from urllib.parse import urlparse

# class to represent a firewall with all the information
class FirewallData:

    def __init__(self,
                 firmware_version: str,
                 license: str,
                 uptime: str,
                 log_disk_pct: str,
                 data_disk_pct: str,
                 HA_status:str,
                 ):
        self.firmware_version = firmware_version
        self.license = license
        self.uptime = uptime
        self.log_disk_pct = log_disk_pct
        self.data_disk_pct = data_disk_pct
        self.HA_status = HA_status


class LoginData:
    
    def __init__(self,firewall,url,username,password):
        self.firewall = firewall
        self.url = url
        self.username = username
        self.password = password
        self.output_dir = ""

coordinate_dic = {
    "firmware_version":"H2",
    "license":"H4",
    "uptime":"H6",
    "log_disk_pct":"H7",
    "data_disk_pct":"H8",
    "HA_status":"H5",
}

def write_report(excel_wb:Workbook,path:pathlib.Path,file_name:str)-> None:
    now = datetime.now()
    date_time = now.strftime("%Y-%m-%d_%H-%M-%S")
    current_date_str = str(date_time)
    file_name = f"{file_name}-{current_date_str}.xlsx"
    path = path.joinpath(file_name)
    try:
        excel_wb.save(path)
    except Exception as e:
        print(f"Unable to save {str(path)} saving to current directory")
        print(e)
        try:
            excel_wb.save(pathlib.Path().cwd().joinpath(file_name))
        except Exception as e:
            print("could not save")
            print(e)

def generate_excel_report(firewall_data:FirewallData) -> Workbook:
    wb = load_workbook("skjema-template.xlsx")
    ws = wb.active

    for key in coordinate_dic:
        coordinate = coordinate_dic[key]
        ws[coordinate] = getattr(firewall_data,key)

    return wb


#find the correct output path for each CSV row.
#the hierarchy goes like
#1: value of --out-dir argument -> 2:value of the csv-file -> 3: current dir
def get_out_dir(row):
    path = pathlib.Path()
    if "--out-dir" in sys.argv:
        return path.joinpath(args.out_dir)
    
    if len(row) == 5:
        return path.joinpath(row[4])
    
    return path.joinpath(os.getcwd())


def get_target_list(str_path:str) -> List[LoginData]:
    ret_list = []
    path = pathlib.Path(str_path)
    try:
        with open(path) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',',quotechar="|")
            for row in csv_reader:
                if len(row) < 4 or len(row) > 5:
                    raise Exception("Bad CSV file")
                url = urlparse(row[1])
                login_data = LoginData(row[0], url, row[2], row[3])
                login_data.output_dir = get_out_dir(row)
                ret_list.append(login_data)

    except FileNotFoundError:
        print(f"Could not find file: {args.target_list}")
    
    except PermissionError:
        print(f"Permission error for file: {args.target_list}")
    
    except Exception as e:
        print("unkown erorr")
        print(e)

    return ret_list

#open a URL
#enter login info and login
#should refresh the page every .get() call
def driver_login(driver,user_data:LoginData):
    driver.get(user_data.url.geturl())
    time.sleep(args.delay)
    title = driver.title
    title_new = ""
    driver.find_element(By.ID, "ELEMENT_login_username").send_keys(user_data.username)
    driver.find_element(By.ID, "ELEMENT_login_password").send_keys(user_data.password)
    driver.find_element(By.ID, "ELEMENT_login_button").click()
    time.sleep(args.delay)

    return driver

"""while title != title_new:
        time.sleep(3)
        title_new = driver.title"""


def get_ha_status(driver):
    #click on the Managment div by xpath
    e = driver.find_element(By.XPATH, "/html/body/div[3]/div[1]/div[1]/div/div/ul/li[2]/div")
    e.click()
    time.sleep(5)

    xpath = "/html/body/div[3]/div[1]/div[1]/div/div/ul/li[2]/ul/li[12]/div"
    e = driver.find_element(By.XPATH,xpath).click()
    time.sleep(5)
    
    xpath = "/html/body/div[3]/div[2]/div/table/tbody/tr[2]/td/div/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr[2]/td/div/div/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[2]/td[2]/table/tbody/tr[1]/td[1]/b"
    status = driver.find_element(By.XPATH,xpath)
    return status.text.split(":")[1].strip()

def get_firmware_version(driver):
    firmware_sibling_node = driver.find_element(
        By.XPATH, "// td[contains(text(),\'Firmware version:')]")
    firmware = firmware_sibling_node.find_element(
        By.XPATH, "./following-sibling::td")
    return firmware.find_element(By.XPATH, "./following-sibling::td").text


#get Exp.Date for Network Protection license
def get_NP_license(driver):
    #click on the Managment div by xpath
    e = driver.find_element(By.XPATH, "/html/body/div[3]/div[1]/div[1]/div/div/ul/li[2]/div")
    e.click()
    time.sleep(5)
    #click on the license tab div
    e = driver.find_element(By.XPATH,"/html/body/div[3]/div[1]/div[1]/div/div/ul/li[2]/ul/li[3]/div").click()
    time.sleep(5)
    
    #retrive the expiration date for the networkprotection date
    e = driver.find_element(By.XPATH, '/html/body/div[3]/div[2]/div/table/tbody/tr[2]/td/div/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr[2]/td/div/div/table/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td[2]/table/tbody/tr[3]/td[3]')
    return e.text


def get_uptime(driver):
    uptime_sibling_node = driver.find_element(
        By.XPATH, "//td[contains(text(),\'Uptime:')]/following-sibling::td")
    uptime = uptime_sibling_node.find_element(
        By.XPATH, "./following-sibling::td")
    return uptime.text


#get the percantage number
#10gb of 20% would return 20% 
def get_percentage(element):
    res = re.findall('\d*%', element.text)
    if len(res) >= 1:
        return res[0]

#get a list with resources in this order [cpu%,ram%,log_disk,data_disk]
def get_resource_usage(driver) -> int:
    classes_oi = "dashboard_usage_bar_txt"
    status = driver.find_elements(By.CLASS_NAME, classes_oi)
    status = map(get_percentage,status)
    #2-4 is only disk usage
    return list(status)[2:4]


def manager_factory(browser_select:str):
    browser_select = browser_select.lower()
    manager_map = {
        "chrome":ChromeDriverManager,
        "firefox":GeckoDriverManager,
        "ie":IEDriverManager,
    }
    return manager_map[browser_select]


def driver_factory(browser_select:str):
    browser_select = browser_select.lower()
    driver_map = {
        "chrome":webdriver.Chrome,
        "firefox":webdriver.Firefox,
        "ie":webdriver.Ie,
    }
    return driver_map[browser_select]


def get_driver(browser_select:str)-> webdriver:
    options = Options()
    options.headless = args.headless

    manager = manager_factory(browser_select)
    driver = driver_factory(browser_select)

    driver = driver(options=options, executable_path=manager().install())
    return driver


#parse bool values from argparse
#https://stackoverflow.com/a/43357954
def str2bool(v):
    if isinstance(v, bool):
        return v
    if v.lower() in ('yes', 'true', 't', 'y', '1'):
        return True
    elif v.lower() in ('no', 'false', 'f', 'n', '0'):
        return False
    else:
        raise argparse.ArgumentTypeError('Boolean value expected.')

arg_epilog="""
\n\n
The format for the TARGET-LIST must be a CSV file with a single comma ',' as a delimiter.\n
'|' can be used as a quotechar.The quotechar should be used for the password field\n
firewall_name,URL,username,password,output-path(optional)\n
"""




def main():


    banner = """       _                 _____           _ _               
      | |               / ____|         (_| |              
      | | ___ _ __  ___| (___   ___ _ __ _| |__   ___ _ __ 
  _   | |/ _ | '_ \/ __|\___ \ / __| '__| | '_ \ / _ | '__|
 | |__| |  __| | | \__ \____) | (__| |  | | |_) |  __| |   
  \____/ \___|_| |_|___|_____/ \___|_|  |_|_.__/ \___|_|   
                                                           
                                                           """
    print(banner)
    print("JensScriber v-1.0 release 2022-1-06")


    arg_parser = argparse.ArgumentParser(
    allow_abbrev=False, description="Retrive information from sopohs firewalls, and save it as a excel report.", formatter_class=argparse.ArgumentDefaultsHelpFormatter,epilog=arg_epilog)

    arg_parser.add_argument("--target-list", required=True, type=str,
                            help="CSV file of firewalls and authentication.")

    arg_parser.add_argument("--browser", required=True, type=str,
                            help="Select which browser to use.(firefox,chrome,IE)",metavar="firefox,chrome,IE")

    arg_parser.add_argument("--out-dir", required=False, default=os.getcwd(), type=str,
                            help="output directory for the reports. This will be used if no output directory is listed in the csv file")

    arg_parser.add_argument("--delay", required=False, default=10, type=int,
                            help="Delay between selenium actions.If the script failts to retrive information, it might be because the page was not loaded in time so it could help to increase the delay value")

    arg_parser.add_argument("--headless",metavar="True/False", required=False, default=True,
                            type=str2bool, help="Should selenium run in headless mode? Useful for debugging")

    global args
    args = arg_parser.parse_args()

    login_list = get_target_list(args.target_list)
    driver = get_driver(args.browser)
    p_bar = tqdm(total=len(login_list * 7),desc="Getting started")

    for login in login_list:
        p_bar.desc = f"Working on {login.firewall}"
        firewall_data = FirewallData
        driver = driver_login(driver,login)
        time.sleep(args.delay)
        
        r = get_resource_usage(driver)
        firewall_data.log_disk_pct = r[0]
        firewall_data.data_disk_pct = r[1]
        p_bar.update(1)
        time.sleep(args.delay)

        firewall_data.firmware_version = get_firmware_version(driver)
        p_bar.update(1)
        time.sleep(args.delay)

        firewall_data.uptime = get_uptime(driver)
        p_bar.update(1)
        time.sleep(args.delay)

        firewall_data.license = get_NP_license(driver)
        p_bar.update(1)
        time.sleep(args.delay)

        firewall_data.HA_status = get_ha_status(driver)
        p_bar.update(1)

        excel_wb = generate_excel_report(firewall_data)
        write_report(excel_wb,login.output_dir,login.firewall)
        p_bar.update(1)


    driver.quit()
    p_bar.close()
    print("DONE")

if __name__ == "__main__":
    main()
